"""
The Annotation Zone
    Jamie Brandon
    Dan Noar
    Irina Onoprienko
    Will Tietz

Script to put compare tags from 3 annotators in an excel spreadsheet

To Do:
"""
    
from openpyxl import Workbook
import xml.etree.ElementTree as ET
from collections import defaultdict
import re

class XML_comp():
    def __init__(self):
        self.sents = []
        self.sent_nums = []
        self.hum_sent_nums = []     # only human tagged sentences
        self.a1_sns = []            # what annoatator 1 tagged
        self.a2_sns = []            # what annoatator 2 tagged
        # main tags = 'NARRATION_AND_DESCRIPTION', 'ABOUT_THE_GAME', 'MECHANICS', 'NON-GAME_RELATED', 'NON-CONTENT'
        # main types is subtags from main tags
        # main is both tag and subtag
        # dec main = descision making or not
        # dec type is subtag for decision making
        # decision is both decision making and its subtag
        dict_types = ['main tag', 'main type', 'main', 'question', 'decision', 'dec main', 'dec type', 'renege']
        self.main_tags = ['NARRATION_AND_DESCRIPTION', 'ABOUT_THE_GAME', 'MECHANICS', 'NON-GAME_RELATED', 'NON-CONTENT']
        self.all_tags = ['NARRATION_AND_DESCRIPTION: description', 'NARRATION_AND_DESCRIPTION: events', 'NARRATION_AND_DESCRIPTION: actions', 'NARRATION_AND_DESCRIPTION: lore_history_backstory', 'ABOUT_THE_GAME: comments', 'ABOUT_THE_GAME: retcon', 'ABOUT_THE_GAME: recap', 'MECHANICS', 'NON-GAME_RELATED', 'NON-CONTENT']
        self.annot1 = {}
        self.annot2 = {}
        self.annot3 = {}
        self.annot4 = {}
        self.annot1_freq = {}
        self.annot2_freq = {}
        self.annot3_freq = {}
        self.annot4_freq = {}
        self.agree = {} #pariwise agreement counter
        # set values for each tag type
        # for the annotator dictionaries it is a default dict with sent# (eg 's42') and tag as value
        for name in dict_types:
            self.annot1[name] = defaultdict(lambda: " ")
            self.annot2[name] = defaultdict(lambda: " ")
            self.annot3[name] = defaultdict(lambda: " ")
            self.annot4[name] = defaultdict(lambda: " ")
            self.agree[name] = 0
        for tag in self.all_tags:
            self.annot1_freq[tag] = 0
            self.annot2_freq[tag] = 0
            self.annot3_freq[tag] = 0
            self.annot4_freq[tag] = 0
        self.wb = Workbook()
    def find_data1(self, file_name):
        """
        Reads an xml file and puts its data in annot1
        Also determines builds a list of sentences
        """
        # read xml file as element tree
        f = ET.parse(file_name)
        root = f.getroot()
        text = root[0]      #everything between <TEXT> </TEXT>
        tags = root[1]      #everything between <TAGS> </TAGS>
        # make list of sentences
        self.sents = text.text.split('\n')
        # each child is a tag
        # child.tag is the name of the tag, eg "DECISION_MAKING" or "ABOUT_THE_GAME"
        # the 'text' attribute is the tagged text (hopefully the sentence number)
        # the 'type' attribute is the subtag
        for child in tags:
            if child.tag == 'DECISION_MAKING':
                self.annot1['decision'][child.attrib['text']] = child.tag + ': ' + child.attrib['type']
                self.annot1['dec main'][child.attrib['text']] = child.tag
                self.annot1['dec type'][child.attrib['text']] = child.attrib['type']
            elif child.tag == 'RENEGE':
                self.annot1['renege'][child.attrib['retconText']] = child.attrib['fake_actionText']
            else:
                self.annot1['main tag'][child.attrib['text']] = child.tag
                # these things don't always exist, so don't add them if they don't
                try:
                    self.annot1['question'][child.attrib['text']] = child.attrib['question']
                    self.annot1['main type'][child.attrib['text']] = child.attrib['type']
                except KeyError:
                    pass
        for sent in self.annot1['main tag']:
            if self.annot1['main tag'][sent] == 'NARRATION_AND_DESCRIPTION' or self.annot1['main tag'][sent] == 'ABOUT_THE_GAME':
                self.annot1['main'][sent] = self.annot1['main tag'][sent] + ': ' + self.annot1['main type'][sent]
                self.annot1_freq[self.annot1['main'][sent]] += 1
            else:
                self.annot1['main'][sent] = self.annot1['main tag'][sent]
                if self.annot1['main tag'][sent] != 'IN-CHARACTER_DIALOGUE' and self.annot1['main tag'][sent] != 'STAGE_DIRECTIONS':
                    self.annot1_freq[self.annot1['main'][sent]] += 1
        # add retconed sentence to retcon subtag
        for sent in self.annot1['main type']:
            if self.annot1['main type'][sent] == 'retcon':
                try:
                    self.annot1['main type'][sent] += ": " + re.findall(r's.*$', self.annot1['renege'][sent])[0]
                except:
                    pass
        # sorted list of sentence numbers
        self.sent_nums = sorted([int(x[1:]) for x in self.annot1['main tag'].keys()])
        # and narrow down on human tagged things
        self.hum_sent_nums = sorted([int(x[1:]) for x in self.annot1['main tag'].keys() if self.annot1['main tag'][x] in self.main_tags])
    def find_data2(self, file_name):
        """
        Reads an xml file and puts its data in annot2
        See find_data1 for comments
        """
        f = ET.parse(file_name)
        root = f.getroot()
        tags = root[1]
        for child in tags:
            if child.tag == 'DECISION_MAKING':
                self.annot2['decision'][child.attrib['text']] = child.tag + ': ' + child.attrib['type']
                self.annot2['dec main'][child.attrib['text']] = child.tag
                self.annot2['dec type'][child.attrib['text']] = child.attrib['type']
            elif child.tag == 'RENEGE':
                self.annot2['renege'][child.attrib['retconText']] = child.attrib['fake_actionText']
            else:
                self.annot2['main tag'][child.attrib['text']] = child.tag
                try:
                    self.annot2['question'][child.attrib['text']] = child.attrib['question']
                    self.annot2['main type'][child.attrib['text']] = child.attrib['type']
                except:
                    pass
        for sent in self.annot2['main tag']:
            if self.annot2['main tag'][sent] == 'NARRATION_AND_DESCRIPTION' or self.annot2['main tag'][sent] == 'ABOUT_THE_GAME':
                self.annot2['main'][sent] = self.annot2['main tag'][sent] + ': ' + self.annot2['main type'][sent]
                if int(sent[1:]) in self.hum_sent_nums:
                    self.annot2_freq[self.annot2['main'][sent]] += 1
            else:
                self.annot2['main'][sent] = self.annot2['main tag'][sent]
                if self.annot2['main tag'][sent] != 'IN-CHARACTER_DIALOGUE' and self.annot2['main tag'][sent] != 'STAGE_DIRECTIONS':
                    if int(sent[1:]) in self.hum_sent_nums:
                        self.annot2_freq[self.annot2['main'][sent]] += 1
        for sent in self.annot2['main type']:
            if self.annot2['main type'][sent] == 'retcon':
                try:
                    self.annot2['main type'][sent] += ": " + re.findall(r's.*$', self.annot2['renege'][sent])[0]
                except:
                    pass
    def find_data3(self, file_name):
        """
        Reads an xml file and puts its data in annot3
        See find_data1 for comments
        """
        f = ET.parse(file_name)
        root = f.getroot()
        tags = root[1]
        for child in tags:
            if child.tag == 'DECISION_MAKING':
                self.annot3['decision'][child.attrib['text']] = child.tag + ': ' + child.attrib['type']
                self.annot3['dec main'][child.attrib['text']] = child.tag
                self.annot3['dec type'][child.attrib['text']] = child.attrib['type']
            elif child.tag == 'RENEGE':
                self.annot3['renege'][child.attrib['retconText']] = child.attrib['fake_actionText']
            else:
                self.annot3['main tag'][child.attrib['text']] = child.tag
                try:
                    self.annot3['question'][child.attrib['text']] = child.attrib['question']
                    self.annot3['main type'][child.attrib['text']] = child.attrib['type']
                except:
                    pass
        for sent in self.annot3['main tag']:
            if self.annot3['main tag'][sent] == 'NARRATION_AND_DESCRIPTION' or self.annot3['main tag'][sent] == 'ABOUT_THE_GAME':
                self.annot3['main'][sent] = self.annot3['main tag'][sent] + ': ' + self.annot3['main type'][sent]
                if int(sent[1:]) in self.hum_sent_nums:
                    self.annot3_freq[self.annot3['main'][sent]] += 1
            else:
                self.annot3['main'][sent] = self.annot3['main tag'][sent]
                if self.annot3['main tag'][sent] != 'IN-CHARACTER_DIALOGUE' and self.annot3['main tag'][sent] != 'STAGE_DIRECTIONS':
                    if int(sent[1:]) in self.hum_sent_nums:
                        self.annot3_freq[self.annot3['main'][sent]] += 1
        for sent in self.annot3['main type']:
            if self.annot3['main type'][sent] == 'retcon':
                try:
                    self.annot3['main type'][sent] += ": " + re.findall(r's.*$', self.annot3['renege'][sent])[0]
                except:
                    pass
    def find_data4(self, file_name):
        """
        Reads an xml file and puts its data in annot3
        See find_data1 for comments
        """
        f = ET.parse(file_name)
        root = f.getroot()
        tags = root[1]
        for child in tags:
            if child.tag == 'DECISION_MAKING':
                self.annot4['decision'][child.attrib['text']] = child.tag + ': ' + child.attrib['type']
                self.annot4['dec main'][child.attrib['text']] = child.tag
                self.annot4['dec type'][child.attrib['text']] = child.attrib['type']
            elif child.tag == 'RENEGE':
                self.annot4['renege'][child.attrib['retconText']] = child.attrib['fake_actionText']
            else:
                self.annot4['main tag'][child.attrib['text']] = child.tag
                try:
                    self.annot4['question'][child.attrib['text']] = child.attrib['question']
                    self.annot4['main type'][child.attrib['text']] = child.attrib['type']
                except:
                    pass
        for sent in self.annot4['main tag']:
            if self.annot4['main tag'][sent] == 'NARRATION_AND_DESCRIPTION' or self.annot4['main tag'][sent] == 'ABOUT_THE_GAME':
                self.annot4['main'][sent] = self.annot4['main tag'][sent] + ': ' + self.annot4['main type'][sent]
                self.annot4_freq[self.annot4['main'][sent]] += 1
            else:
                self.annot4['main'][sent] = self.annot4['main tag'][sent]
                if self.annot4['main tag'][sent] != 'IN-CHARACTER_DIALOGUE' and self.annot4['main tag'][sent] != 'STAGE_DIRECTIONS':
                    self.annot4_freq[self.annot4['main'][sent]] += 1
        for sent in self.annot4['main type']:
            if self.annot4['main type'][sent] == 'retcon':
                self.annot4['main type'][sent] += ": " + re.findall(r's.*$', self.annot4['renege'][sent])[0]
    def color(self, row, sent, ws, ag_col):
        """
        colors everything, uses color_helper to check agreement
        """
        self.color_helper('main tag', 2, row, sent, ws)
        self.color_helper('main type', 3, row, sent, ws)
        self.color_helper('question', 4, row, sent, ws)
        if self.agree_check(sent) == "yes":
            ws.cell(row=row, column=ag_col).style = 'Good'
        else:
            ws.cell(row=row, column=ag_col).style = 'Bad'
    def table_1_2(self, ws):
        """
        prints contingency table between annotator 1 and 2 for main tag
        """
        matrix = defaultdict(int)
        for sent in self.hum_sent_nums:
            if self.annot1['main tag']['s'+str(sent)] != 'IN-CHARACTER_DIALOGUE' and self.annot1['main tag']['s'+str(sent)] != 'STAGE_DIRECTIONS':
                matrix[(self.annot1['main tag']['s'+str(sent)], self.annot2['main tag']['s'+str(sent)])] += 1
        ws.cell(row=1, column=2, value='NARRATION_AND_DESCRIPTION')
        ws.cell(row=1, column=3, value='ABOUT_THE_GAME')
        ws.cell(row=1, column=4, value='MECHANICS')
        ws.cell(row=1, column=5, value='NON-GAME_RELATED')
        ws.cell(row=1, column=6, value='NON-CONTENT')
        ws.cell(row=2, column=1, value='NARRATION_AND_DESCRIPTION')
        ws.cell(row=3, column=1, value='ABOUT_THE_GAME')
        ws.cell(row=4, column=1, value='MECHANICS')
        ws.cell(row=5, column=1, value='NON-GAME_RELATED')
        ws.cell(row=6, column=1, value='NON-CONTENT')
        rows = 2
        cols = 2
        for tag1 in self.main_tags:
            for tag2 in self.main_tags:
                ws.cell(row=rows, column=cols, value=matrix[(tag1, tag2)])
                cols += 1
            rows += 1
            cols = 2
    def tag_counter(self, ws, start_row, column, letter):
        """
        prints an excel formula to count tags
        """
        rows = start_row
        for tag in self.main_tags:
            ws.cell(row=rows, column=column, value='=COUNTIF(Data!' + letter + ':' + letter + ', "' + tag + '")')
            rows += 1
    def table_1_2_all(self, ws):
        """
        prints contingency table between annotator 1 and 2 for everything
        """
        matrix = defaultdict(int)
        for sent in self.hum_sent_nums:
            if self.annot1['main']['s'+str(sent)] != 'IN-CHARACTER_DIALOGUE' and self.annot1['main']['s'+str(sent)] != 'STAGE_DIRECTIONS':
                matrix[(self.annot1['main']['s'+str(sent)], self.annot2['main']['s'+str(sent)])] += 1
        rows = 2
        cols = 11
        for tag in self.all_tags:
            ws.cell(row=1, column=cols, value=tag)
            ws.cell(row=rows, column=10, value=tag)
            rows += 1
            cols += 1
        rows = 2
        cols = 11
        for tag1 in self.all_tags:
            for tag2 in self.all_tags:
                ws.cell(row=rows, column=cols, value=matrix[(tag1, tag2)])
                cols += 1
            rows += 1
            cols = 11
    
class two_annots(XML_comp):
    def print_to_excel(self, file_name):
        ws = self.wb.active
        ws.title = "Data"
        # columns headers
        ws.cell(row=1, column=3, value="annotator 1")
        ws.cell(row=1, column=6, value="annotator 2")
        ws.cell(row=2, column=1, value="sent#")
        ws.cell(row=2, column=2, value="tag")
        ws.cell(row=2, column=3, value="type")
        ws.cell(row=2, column=4, value="question")
        ws.cell(row=2, column=5, value="tag")
        ws.cell(row=2, column=6, value="type")
        ws.cell(row=2, column=7, value="question")
        ws.cell(row=2, column=8, value="agree")
        ws.cell(row=2, column=9, value="sent")
        row = 3
        # print each human tagged sentence with its tags
        for num in self.hum_sent_nums:
            sent = "s" + str(num)
            if self.annot1['main tag'][sent] != 'IN-CHARACTER_DIALOGUE' and self.annot1['main tag'][sent] != 'STAGE_DIRECTIONS':
                ws.cell(row=row, column=1, value=sent)
                ws.cell(row=row, column=2, value=self.annot1['main tag'][sent])
                ws.cell(row=row, column=3, value=self.annot1['main type'][sent])
                ws.cell(row=row, column=4, value=self.annot1['question'][sent])
                ws.cell(row=row, column=5, value=self.annot2['main tag'][sent])
                ws.cell(row=row, column=6, value=self.annot2['main type'][sent])
                ws.cell(row=row, column=7, value=self.annot2['question'][sent])
                ws.cell(row=row, column=8, value=self.agree_check(sent))
                ws.cell(row=row, column=9, value=self.sents[num+1])
                # call self.color to color everything
                self.color(row, sent, ws, 8)
                row += 1
        self.wb.save(file_name)
    def agree_check(self, sent):
        """
        checks if there is universal agreement for every tag
        returns yes or no since that is what should be printed in excel
        """
        if self.annot1['main tag'][sent] == self.annot2['main tag'][sent] and \
        self.annot1['main type'][sent] == self.annot2['main type'][sent] and \
        self.annot1['question'][sent] == self.annot2['question'][sent]:
            return "yes"
        else:
            return "no"
    def color_helper(self, tag, col, row, sent, ws):
        """
        determines how much agreement there is and colors the attribute accordingly
        """
        if self.annot1[tag][sent] == self.annot2[tag][sent]:
            ws.cell(row=row, column=col).style = 'Good'
            ws.cell(row=row, column=col+3).style = 'Good'
        else:
            ws.cell(row=row, column=col).style = 'Bad'
            ws.cell(row=row, column=col+3).style = 'Bad'
    def disagree(self):
        length = len(self.hum_sent_nums)
        disagreements = 0
        for i in range(len(self.hum_sent_nums)-15):
            check = self.hum_sent_nums[i:i+15]
            total = 15
            wrong = 0
            for sent in check:
                if self.annot1['main']['s'+str(sent)] != self.annot2['main']['s'+str(sent)]:
                    wrong += 1
            
            if wrong/total >= .7:
                j = i + 1
                while j < len(self.hum_sent_nums):
                    check.append(self.hum_sent_nums[j])
                    total = len(check)
                    wrong = 0
                    for sent in check:
                        if self.annot1['main']['s'+str(sent)] != self.annot2['main']['s'+str(sent)]:
                            wrong += 1
                    j += 1
                    if wrong/total < .7:
                        break
                check.reverse()
                disagreements += wrong
                for sent in check:
                    if self.annot1['main']['s'+str(sent)] == self.annot2['main']['s'+str(sent)]:
                        check.remove(sent)
                    else:
                        break
                print(length-len(self.hum_sent_nums))
        return disagreements
                    
    def stats(self, file_name):
        ws = self.wb.create_sheet("tables")
        self.table_1_2(ws)
        self.table_1_2_all(ws)
        ws.cell(row=12, column=1, value='NARRATION_AND_DESCRIPTION')
        ws.cell(row=13, column=1, value='ABOUT_THE_GAME')
        ws.cell(row=14, column=1, value='MECHANICS')
        ws.cell(row=15, column=1, value='NON-GAME_RELATED')
        ws.cell(row=16, column=1, value='NON-CONTENT')
        ws.cell(row=11, column=2, value='annot 1')
        self.tag_counter(ws, 12, 2, 'B')
        ws.cell(row=11, column=3, value='annot 2')
        self.tag_counter(ws, 12, 3, 'E')
        ws.cell(row=11, column=4, value='agreement')
        ws.cell(row=12, column=4, value='=B2')
        ws.cell(row=13, column=4, value='=C3')
        ws.cell(row=14, column=4, value='=D4')
        ws.cell(row=15, column=4, value='=E5')
        ws.cell(row=16, column=4, value='=F6')
        ws.cell(row=17, column=2, value='=sum(B12:B16)')
        ws.cell(row=17, column=3, value='=sum(C12:C16)')
        ws.cell(row=17, column=4, value='=sum(D12:D16)')
        ws.cell(row=18, column=4, value='=D17/B17')
        ws.cell(row=12, column=5, value='=B12/B17')
        ws.cell(row=13, column=5, value='=B13/B17')
        ws.cell(row=14, column=5, value='=B14/B17')
        ws.cell(row=15, column=5, value='=B15/B17')
        ws.cell(row=16, column=5, value='=B16/B17')
        ws.cell(row=12, column=6, value='=C12/C17')
        ws.cell(row=13, column=6, value='=C13/C17')
        ws.cell(row=14, column=6, value='=C14/C17')
        ws.cell(row=15, column=6, value='=C15/C17')
        ws.cell(row=16, column=6, value='=C16/C17')
        ws.cell(row=12, column=7, value='=E12*F12')
        ws.cell(row=13, column=7, value='=E13*F13')
        ws.cell(row=14, column=7, value='=E14*F14')
        ws.cell(row=15, column=7, value='=E15*F15')
        ws.cell(row=16, column=7, value='=E16*F16')
        ws.cell(row=17, column=7, value='=sum(G12:G16)')
        ws.cell(row=19, column=5, value='κ')
        ws.cell(row=19, column=6, value='=(D18-G17)/(1-G17)')
        
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=10, value=tag)
            rows += 1
        
        ws.cell(row=20, column=11, value='annot 1')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=11, value=self.annot1_freq[tag])
            rows += 1
        
        ws.cell(row=20, column=12, value='annot 2')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=12, value=self.annot2_freq[tag])
            rows += 1
        
        ws.cell(row=20, column=13, value='agreement')
        ws.cell(row=21, column=13, value='=K2')
        ws.cell(row=22, column=13, value='=L3')
        ws.cell(row=23, column=13, value='=M4')
        ws.cell(row=24, column=13, value='=N5')
        ws.cell(row=25, column=13, value='=O6')
        ws.cell(row=26, column=13, value='=P7')
        ws.cell(row=27, column=13, value='=Q8')
        ws.cell(row=28, column=13, value='=R9')
        ws.cell(row=29, column=13, value='=S10')
        ws.cell(row=30, column=13, value='=T11')
        
        ws.cell(row=31, column=11, value='=sum(K21:K30)')
        ws.cell(row=31, column=12, value='=sum(L21:L30)')
        ws.cell(row=31, column=13, value='=sum(M21:M30)')
        ws.cell(row=32, column=13, value='=M31/K31')
        
        ws.cell(row=20, column=14, value='P(tag|a1)')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=14, value='=K'+str(rows)+'/'+str(len(self.hum_sent_nums)))
            rows += 1
            
        ws.cell(row=20, column=15, value='P(tag|a2)')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=15, value='=L'+str(rows)+'/'+str(len(self.hum_sent_nums)))
            rows += 1
        
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=16, value='=N'+str(rows)+'*'+'O'+str(rows))
            rows += 1
        ws.cell(row=rows, column=16, value='=sum(P21:P30)')
        
        ws.cell(row=33, column=14, value='κ')
        ws.cell(row=33, column=15, value='=(M32-P31)/(1-P31)')
        
        ws.cell(row=15, column=12, value='predicted:')
        ws.cell(row=16, column=12, value='observed:')
        ws.cell(row=17, column=12, value='κ')
        ws.cell(row=15, column=13, value='=0.7*G17+0.3*P31')
        ws.cell(row=16, column=13, value='=0.7*D18+0.3*M32')
        ws.cell(row=17, column=13, value='=(M16-M15)/(1-M15)')
        
        self.wb.save(file_name)
        
class gold(XML_comp):
    """
    this class is using information from the gold standard put in annot1
    so annotator 1 is in annot2 and annotator 2 is in annot3
    """
    def print_to_excel(self, file_name):
        ws = self.wb.active
        ws.title = "Data"
        # columns headers
        ws.cell(row=1, column=3, value="annotator 1")
        ws.cell(row=1, column=6, value="annotator 2")
        ws.cell(row=2, column=1, value="sent#")
        ws.cell(row=2, column=2, value="tag")
        ws.cell(row=2, column=3, value="type")
        ws.cell(row=2, column=4, value="question")
        ws.cell(row=2, column=5, value="tag")
        ws.cell(row=2, column=6, value="type")
        ws.cell(row=2, column=7, value="question")
        ws.cell(row=2, column=8, value="agree")
        ws.cell(row=2, column=9, value="sent")
        row = 3
        # print each human tagged sentence with its tags
        for num in self.hum_sent_nums:
            sent = "s" + str(num)
            ws.cell(row=row, column=1, value=sent)
            ws.cell(row=row, column=2, value=self.annot2['main tag'][sent])
            ws.cell(row=row, column=3, value=self.annot2['main type'][sent])
            ws.cell(row=row, column=4, value=self.annot2['question'][sent])
            ws.cell(row=row, column=5, value=self.annot3['main tag'][sent])
            ws.cell(row=row, column=6, value=self.annot3['main type'][sent])
            ws.cell(row=row, column=7, value=self.annot3['question'][sent])
            ws.cell(row=row, column=8, value=self.agree_check(sent))
            ws.cell(row=row, column=9, value=self.sents[num+1])
            # call self.color to color everything
            self.color(row, sent, ws, 8)
            row += 1
        self.wb.save(file_name)
        
    def agree_check(self, sent):
        """
        checks if there is universal agreement for every tag
        returns yes or no since that is what should be printed in excel
        """
        if self.annot2['main tag'][sent] == self.annot3['main tag'][sent] and \
        self.annot2['main type'][sent] == self.annot3['main type'][sent] and \
        self.annot2['question'][sent] == self.annot3['question'][sent]:
            return "yes"
        else:
            return "no"
        
    def color_helper(self, tag, col, row, sent, ws):
        """
        determines how much agreement there is and colors the attribute accordingly
        """
        if self.annot2[tag][sent] == self.annot3[tag][sent]:
            ws.cell(row=row, column=col).style = 'Good'
            ws.cell(row=row, column=col+3).style = 'Good'
        else:
            ws.cell(row=row, column=col).style = 'Bad'
            ws.cell(row=row, column=col+3).style = 'Bad'
            
    def table_1_2(self, ws):
        """
        prints contingency table between annotator 1 and 2 for main tag
        """
        matrix = defaultdict(int)
        for sent in self.hum_sent_nums:
            matrix[(self.annot2['main tag']['s'+str(sent)], self.annot3['main tag']['s'+str(sent)])] += 1
        ws.cell(row=1, column=2, value='NARRATION_AND_DESCRIPTION')
        ws.cell(row=1, column=3, value='ABOUT_THE_GAME')
        ws.cell(row=1, column=4, value='MECHANICS')
        ws.cell(row=1, column=5, value='NON-GAME_RELATED')
        ws.cell(row=1, column=6, value='NON-CONTENT')
        ws.cell(row=2, column=1, value='NARRATION_AND_DESCRIPTION')
        ws.cell(row=3, column=1, value='ABOUT_THE_GAME')
        ws.cell(row=4, column=1, value='MECHANICS')
        ws.cell(row=5, column=1, value='NON-GAME_RELATED')
        ws.cell(row=6, column=1, value='NON-CONTENT')
        rows = 2
        cols = 2
        for tag1 in self.main_tags:
            for tag2 in self.main_tags:
                ws.cell(row=rows, column=cols, value=matrix[(tag1, tag2)])
                cols += 1
            rows += 1
            cols = 2
    
    def table_1_2_all(self, ws):
        """
        prints contingency table between annotator 1 and 2 for everything
        """
        matrix = defaultdict(int)
        for sent in self.hum_sent_nums:
            matrix[(self.annot2['main']['s'+str(sent)], self.annot3['main']['s'+str(sent)])] += 1
        print(matrix)
        rows = 2
        cols = 11
        for tag in self.all_tags:
            ws.cell(row=1, column=cols, value=tag)
            ws.cell(row=rows, column=10, value=tag)
            rows += 1
            cols += 1
        rows = 2
        cols = 11
        for tag1 in self.all_tags:
            for tag2 in self.all_tags:
                ws.cell(row=rows, column=cols, value=matrix[(tag1, tag2)])
                cols += 1
            rows += 1
            cols = 11
    
    def disagree(self):
        length = len(self.hum_sent_nums)
        disagreements = 0
        overall = 0
        i = 0
        while i < (len(self.hum_sent_nums)-15):
            check = self.hum_sent_nums[i:i+15]
            total = 15
            wrong = 0
            for sent in check:
                if self.annot2['main']['s'+str(sent)] != self.annot3['main']['s'+str(sent)]:
                    wrong += 1
            
            if wrong/total >= .7:
                j = i + 1
                while j < len(self.hum_sent_nums):
                    check.append(self.hum_sent_nums[j])
                    total = len(check)
                    wrong = 0
                    for sent in check:
                        if self.annot2['main']['s'+str(sent)] != self.annot3['main']['s'+str(sent)]:
                            wrong += 1
                    j += 1
                    if wrong/total < .7:
                        break
                check.reverse()
                for sent in check:
                    if self.annot2['main']['s'+str(sent)] == self.annot3['main']['s'+str(sent)]:
                        check.remove(sent)
                        j -= 1
                    else:
                        break
                disagreements += wrong
                overall += total
                i = j
            i += 1
        return disagreements, overall
    
    def question_matrix(self, ws):
        matrix = defaultdict(int)
        for sent in self.hum_sent_nums:
            matrix[(self.annot2['question']['s'+str(sent)], self.annot3['question']['s'+str(sent)])] += 1
        ws.cell(row=40, column=2, value='YES')
        ws.cell(row=40, column=3, value='NO')
        ws.cell(row=40, column=4, value='N/A')
        ws.cell(row=41, column=1, value='YES')
        ws.cell(row=42, column=1, value='NO')
        ws.cell(row=43, column=1, value='N/A')
        rows = 41
        cols = 2
        q_types = ['YES', 'NO', ' ']
        for tag1 in q_types:
            for tag2 in q_types:
                ws.cell(row=rows, column=cols, value=matrix[(tag1, tag2)])
                cols += 1
            rows += 1
            cols = 2
            
    
    def stats(self, file_name):
        ws = self.wb.create_sheet("tables")
        self.table_1_2(ws)
        self.table_1_2_all(ws)
        # main tag IAA chart
        ws.cell(row=12, column=1, value='NARRATION_AND_DESCRIPTION')
        ws.cell(row=13, column=1, value='ABOUT_THE_GAME')
        ws.cell(row=14, column=1, value='MECHANICS')
        ws.cell(row=15, column=1, value='NON-GAME_RELATED')
        ws.cell(row=16, column=1, value='NON-CONTENT')
        # count tag usage
        ws.cell(row=11, column=2, value='annot 1')
        self.tag_counter(ws, 12, 2, 'B')
        ws.cell(row=11, column=3, value='annot 2')
        self.tag_counter(ws, 12, 3, 'E')
        # count agreement from matrix
        ws.cell(row=11, column=4, value='agreement')
        ws.cell(row=12, column=4, value='=B2')
        ws.cell(row=13, column=4, value='=C3')
        ws.cell(row=14, column=4, value='=D4')
        ws.cell(row=15, column=4, value='=E5')
        ws.cell(row=16, column=4, value='=F6')
        # calculate stats
        ws.cell(row=17, column=2, value='=sum(B12:B16)')
        ws.cell(row=17, column=3, value='=sum(C12:C16)')
        ws.cell(row=17, column=4, value='=sum(D12:D16)')
        ws.cell(row=18, column=4, value='=D17/B17')
        ws.cell(row=12, column=5, value='=B12/B17')
        ws.cell(row=13, column=5, value='=B13/B17')
        ws.cell(row=14, column=5, value='=B14/B17')
        ws.cell(row=15, column=5, value='=B15/B17')
        ws.cell(row=16, column=5, value='=B16/B17')
        ws.cell(row=12, column=6, value='=C12/C17')
        ws.cell(row=13, column=6, value='=C13/C17')
        ws.cell(row=14, column=6, value='=C14/C17')
        ws.cell(row=15, column=6, value='=C15/C17')
        ws.cell(row=16, column=6, value='=C16/C17')
        ws.cell(row=12, column=7, value='=E12*F12')
        ws.cell(row=13, column=7, value='=E13*F13')
        ws.cell(row=14, column=7, value='=E14*F14')
        ws.cell(row=15, column=7, value='=E15*F15')
        ws.cell(row=16, column=7, value='=E16*F16')
        ws.cell(row=17, column=7, value='=sum(G12:G16)')
        ws.cell(row=19, column=5, value='κ')
        ws.cell(row=19, column=6, value='=(D18-G17)/(1-G17)')
        
        # build full tag agreement chart
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=10, value=tag)
            rows += 1
        # print tag usage
        ws.cell(row=20, column=11, value='annot 1')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=11, value=self.annot2_freq[tag])
            rows += 1
        
        ws.cell(row=20, column=12, value='annot 2')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=12, value=self.annot3_freq[tag])
            rows += 1
        # get agreement from matrix
        ws.cell(row=20, column=13, value='agreement')
        ws.cell(row=21, column=13, value='=K2')
        ws.cell(row=22, column=13, value='=L3')
        ws.cell(row=23, column=13, value='=M4')
        ws.cell(row=24, column=13, value='=N5')
        ws.cell(row=25, column=13, value='=O6')
        ws.cell(row=26, column=13, value='=P7')
        ws.cell(row=27, column=13, value='=Q8')
        ws.cell(row=28, column=13, value='=R9')
        ws.cell(row=29, column=13, value='=S10')
        ws.cell(row=30, column=13, value='=T11')
        # calculate stats
        ws.cell(row=31, column=11, value='=sum(K21:K30)')
        ws.cell(row=31, column=12, value='=sum(L21:L30)')
        ws.cell(row=31, column=13, value='=sum(M21:M30)')
        ws.cell(row=32, column=13, value='=M31/K31')
        
        ws.cell(row=20, column=14, value='P(tag|a1)')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=14, value='=K'+str(rows)+'/'+str(len(self.hum_sent_nums)))
            rows += 1
            
        ws.cell(row=20, column=15, value='P(tag|a2)')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=15, value='=L'+str(rows)+'/'+str(len(self.hum_sent_nums)))
            rows += 1
        
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=16, value='=N'+str(rows)+'*'+'O'+str(rows))
            rows += 1
        ws.cell(row=rows, column=16, value='=sum(P21:P30)')
        
        ws.cell(row=33, column=14, value='κ')
        ws.cell(row=33, column=15, value='=(M32-P31)/(1-P31)')
        
        ws.cell(row=15, column=12, value='predicted:')
        ws.cell(row=16, column=12, value='observed:')
        ws.cell(row=17, column=12, value='κ')
        ws.cell(row=15, column=13, value='=0.7*G17+0.3*P31')
        ws.cell(row=16, column=13, value='=0.7*D18+0.3*M32')
        ws.cell(row=17, column=13, value='=(M16-M15)/(1-M15)')
        # disagreement clusters
        dis, overall = self.disagree()
        ws.cell(row=34, column=10, value='total clusters')
        ws.cell(row=34, column=11, value=overall)
        ws.cell(row=34, column=12, value=overall/len(self.hum_sent_nums))
        ws.cell(row=36, column=10, value='disagreement clusters')
        ws.cell(row=36, column=11, value=dis)
        ws.cell(row=36, column=12, value='='+ str(dis)+'/('+str(len(self.hum_sent_nums))+'-M31)')
        
        self.question_matrix(ws)
        
        self.wb.save(file_name)

class three_annots(XML_comp):
    def print_to_excel(self, file_name):
        ws = self.wb.active
        ws.title = "Data"
        # columns headers
        ws.cell(row=1, column=3, value="annotator 1")
        ws.cell(row=1, column=6, value="annotator 2")
        ws.cell(row=1, column=9, value="annotator 3")
        ws.cell(row=2, column=1, value="sent#")
        ws.cell(row=2, column=2, value="tag")
        ws.cell(row=2, column=3, value="type")
        ws.cell(row=2, column=4, value="question")
        ws.cell(row=2, column=5, value="tag")
        ws.cell(row=2, column=6, value="type")
        ws.cell(row=2, column=7, value="question")
        ws.cell(row=2, column=8, value="tag")
        ws.cell(row=2, column=9, value="type")
        ws.cell(row=2, column=10, value="question")
        ws.cell(row=2, column=11, value="agree")
        ws.cell(row=2, column=12, value="sent")
        row = 3
        # print each human tagged sentence with its tags
        for num in self.sent_nums:
            sent = "s" + str(num)
            if self.annot1['main tag'][sent] != 'IN-CHARACTER_DIALOGUE' and self.annot1['main tag'][sent] != 'STAGE_DIRECTIONS':
                ws.cell(row=row, column=1, value=sent)
                ws.cell(row=row, column=2, value=self.annot1['main tag'][sent])
                ws.cell(row=row, column=3, value=self.annot1['main type'][sent])
                ws.cell(row=row, column=4, value=self.annot1['question'][sent])
                ws.cell(row=row, column=5, value=self.annot2['main tag'][sent])
                ws.cell(row=row, column=6, value=self.annot2['main type'][sent])
                ws.cell(row=row, column=7, value=self.annot2['question'][sent])
                ws.cell(row=row, column=8, value=self.annot3['main tag'][sent])
                ws.cell(row=row, column=9, value=self.annot3['main type'][sent])
                ws.cell(row=row, column=10, value=self.annot3['question'][sent])
                ws.cell(row=row, column=12, value=self.agree_check(sent))
                ws.cell(row=row, column=13, value=self.sents[num+1])
                # call self.color to color everything
                self.color(row, sent, ws, 12)
                row += 1
        self.wb.save(file_name)
    def agree_check(self, sent):
        """
        checks if there is universal agreement for every tag
        returns yes or no since that is what should be printed in excel
        """
        if self.annot1['main tag'][sent] == self.annot2['main tag'][sent] == self.annot3['main tag'][sent] and \
        self.annot1['main type'][sent] == self.annot2['main type'][sent] == self.annot3['main type'][sent] and \
        self.annot1['question'][sent] == self.annot2['question'][sent] == self.annot3['question'][sent]:
            return "yes"
        else:
            return "no"
    def color_helper(self, tag_type, col, row, sent, ws):
        """
        determines how much agreement there is and colors the attribute accordingly
        """
        if self.annot1[tag_type][sent] == self.annot2[tag_type][sent] == self.annot3[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Good'
            ws.cell(row=row, column=col+3).style = 'Good'
            ws.cell(row=row, column=col+6).style = 'Good'
            self.agree[tag_type] += 3
        elif self.annot1[tag_type][sent] == self.annot2[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Neutral'
            ws.cell(row=row, column=col+3).style = 'Neutral'
            ws.cell(row=row, column=col+6).style = 'Bad'
            self.agree[tag_type] += 1
        elif self.annot1[tag_type][sent] == self.annot3[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Neutral'
            ws.cell(row=row, column=col+3).style = 'Bad'
            ws.cell(row=row, column=col+6).style = 'Neutral'
            self.agree[tag_type] += 1
        elif self.annot2[tag_type][sent] == self.annot3[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Bad'
            ws.cell(row=row, column=col+3).style = 'Neutral'
            ws.cell(row=row, column=col+6).style = 'Neutral'
            self.agree[tag_type] += 1
        else:
            ws.cell(row=row, column=col).style = 'Bad'
            ws.cell(row=row, column=col+3).style = 'Bad'
            ws.cell(row=row, column=col+6).style = 'Bad'
    def stats(self, file_name):
        ws = self.wb.create_sheet("tables")
        self.table_1_2(ws)
        ws.cell(row=12, column=1, value='NARRATION_AND_DESCRIPTION')
        ws.cell(row=13, column=1, value='ABOUT_THE_GAME')
        ws.cell(row=14, column=1, value='MECHANICS')
        ws.cell(row=15, column=1, value='NON-GAME_RELATED')
        ws.cell(row=16, column=1, value='NON-CONTENT')
        ws.cell(row=11, column=2, value='annot 1')
        self.tag_counter(ws, 12, 2, 'B')
        ws.cell(row=11, column=3, value='annot 2')
        self.tag_counter(ws, 12, 3, 'E')
        ws.cell(row=11, column=4, value='annot 3')
        self.tag_counter(ws, 12, 4, 'H')
        ws.cell(row=11, column=5, value='sum')
        rows = 12
        for tag in self.main_tags:
            ws.cell(row=rows, column=5, value='=sum(B'+str(rows)+':D'+str(rows)+')')
            rows += 1
        ws.cell(row=11, column=6, value='P(tag)')
        rows = 12
        for tag in self.main_tags:
            ws.cell(row=rows, column=6, value='=E'+str(rows)+'/'+str(3*len(self.hum_sent_nums)))
            rows += 1
        ws.cell(row=17, column=5, value='chance:')
        ws.cell(row=17, column=6, value='=F12^2+F13^2+F14^2+F15^2+F16^2')
        ws.cell(row=18, column=5, value='observed')
        ws.cell(row=18, column=6, value=self.agree['main tag']/(3*len(self.hum_sent_nums)))
        ws.cell(row=19, column=5, value='κ')
        ws.cell(row=19, column=6, value='=(F18-F17)/(1-F17)')
        self.wb.save(file_name)

class four_annots(XML_comp):
    def print_to_excel(self, file_name):
        ws = self.wb.active
        ws.title = "Data"
        # columns headers
        ws.cell(row=1, column=3, value="annotator 1")
        ws.cell(row=1, column=6, value="annotator 2")
        ws.cell(row=1, column=9, value="annotator 3")
        ws.cell(row=1, column=12, value="annotator 4")
        ws.cell(row=2, column=1, value="sent#")
        ws.cell(row=2, column=2, value="tag")
        ws.cell(row=2, column=3, value="type")
        ws.cell(row=2, column=4, value="question")
        ws.cell(row=2, column=5, value="tag")
        ws.cell(row=2, column=6, value="type")
        ws.cell(row=2, column=7, value="question")
        ws.cell(row=2, column=8, value="tag")
        ws.cell(row=2, column=9, value="type")
        ws.cell(row=2, column=10, value="question")
        ws.cell(row=2, column=11, value="tag")
        ws.cell(row=2, column=12, value="type")
        ws.cell(row=2, column=13, value="question")
        ws.cell(row=2, column=14, value="agree")
        ws.cell(row=2, column=15, value="sent")
        row = 3
        # print each human tagged sentence with its tags
        for num in self.sent_nums:
            sent = "s" + str(num)
            if self.annot1['main tag'][sent] != 'IN-CHARACTER_DIALOGUE' and self.annot1['main tag'][sent] != 'STAGE_DIRECTIONS':
                ws.cell(row=row, column=1, value=sent)
                ws.cell(row=row, column=2, value=self.annot1['main tag'][sent])
                ws.cell(row=row, column=3, value=self.annot1['main type'][sent])
                ws.cell(row=row, column=4, value=self.annot1['question'][sent])
                ws.cell(row=row, column=5, value=self.annot2['main tag'][sent])
                ws.cell(row=row, column=6, value=self.annot2['main type'][sent])
                ws.cell(row=row, column=7, value=self.annot2['question'][sent])
                ws.cell(row=row, column=8, value=self.annot3['main tag'][sent])
                ws.cell(row=row, column=9, value=self.annot3['main type'][sent])
                ws.cell(row=row, column=10, value=self.annot3['question'][sent])
                ws.cell(row=row, column=11, value=self.annot4['main tag'][sent])
                ws.cell(row=row, column=12, value=self.annot4['main type'][sent])
                ws.cell(row=row, column=13, value=self.annot4['question'][sent])
                ws.cell(row=row, column=14, value=self.agree_check(sent))
                ws.cell(row=row, column=15, value=self.sents[num+1])
                # call self.color to color everything
                self.color(row, sent, ws, 14)
                self.pair_helper('main tag', sent, ws, .7)
                self.pair_helper('main', sent, ws, .3)
                row += 1
        self.wb.save(file_name)
    def agree_check(self, sent):
        """
        checks if there is universal agreement for every tag
        returns yes or no since that is what should be printed in excel
        """
        if self.annot1['main tag'][sent] == self.annot2['main tag'][sent] == self.annot3['main tag'][sent] == self.annot4['main tag'][sent] and \
        self.annot1['main type'][sent] == self.annot2['main type'][sent] == self.annot3['main type'][sent] == self.annot4['main type'][sent] and \
        self.annot1['question'][sent] == self.annot2['question'][sent] == self.annot3['question'][sent] == self.annot4['question'][sent]:
            return "yes"
        else:
            return "no"
    def color_helper(self, tag_type, col, row, sent, ws):
        """
        determines how much agreement there is and colors the attribute accordingly
        """
        if self.annot1[tag_type][sent] == self.annot2[tag_type][sent] == self.annot3[tag_type][sent] == self.annot4[tag_type][sent]:
            ws.cell(row=row, column=col).style = '40 % - Accent5'
            ws.cell(row=row, column=col+3).style = '40 % - Accent5'
            ws.cell(row=row, column=col+6).style = '40 % - Accent5'
            ws.cell(row=row, column=col+9).style = '40 % - Accent5'
            self.agree[tag_type] += 4
        elif self.annot1[tag_type][sent] == self.annot2[tag_type][sent] == self.annot3[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Good'
            ws.cell(row=row, column=col+3).style = 'Good'
            ws.cell(row=row, column=col+6).style = 'Good'
            ws.cell(row=row, column=col+9).style = 'Bad'
            self.agree[tag_type] += 3
        elif self.annot1[tag_type][sent] == self.annot2[tag_type][sent] == self.annot4[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Good'
            ws.cell(row=row, column=col+3).style = 'Good'
            ws.cell(row=row, column=col+6).style = 'Bad'
            ws.cell(row=row, column=col+9).style = 'Good'
            self.agree[tag_type] += 3
        elif self.annot1[tag_type][sent] == self.annot3[tag_type][sent] == self.annot4[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Good'
            ws.cell(row=row, column=col+3).style = 'Bad'
            ws.cell(row=row, column=col+6).style = 'Good'
            ws.cell(row=row, column=col+9).style = 'Good'
            self.agree[tag_type] += 3
        elif self.annot2[tag_type][sent] == self.annot3[tag_type][sent] == self.annot4[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Bad'
            ws.cell(row=row, column=col+3).style = 'Good'
            ws.cell(row=row, column=col+6).style = 'Good'
            ws.cell(row=row, column=col+9).style = 'Good'
            self.agree[tag_type] += 3
        elif self.annot1[tag_type][sent] == self.annot2[tag_type][sent] and self.annot3[tag_type][sent] == self.annot4[tag_type][sent]:
            ws.cell(row=row, column=col).style = '40 % - Accent6'
            ws.cell(row=row, column=col+3).style = '40 % - Accent6'
            ws.cell(row=row, column=col+6).style = '40 % - Accent4'
            ws.cell(row=row, column=col+9).style = '40 % - Accent4'
            self.agree[tag_type] += 2
        elif self.annot1[tag_type][sent] == self.annot3[tag_type][sent] and self.annot2[tag_type][sent] == self.annot4[tag_type][sent]:
            ws.cell(row=row, column=col).style = '40 % - Accent6'
            ws.cell(row=row, column=col+3).style = '40 % - Accent4'
            ws.cell(row=row, column=col+6).style = '40 % - Accent6'
            ws.cell(row=row, column=col+9).style = '40 % - Accent4'
            self.agree[tag_type] += 2
        elif self.annot1[tag_type][sent] == self.annot4[tag_type][sent] and self.annot2[tag_type][sent] == self.annot3[tag_type][sent]:
            ws.cell(row=row, column=col).style = '40 % - Accent6'
            ws.cell(row=row, column=col+3).style = '40 % - Accent4'
            ws.cell(row=row, column=col+6).style = '40 % - Accent4'
            ws.cell(row=row, column=col+9).style = '40 % - Accent6'
            self.agree[tag_type] += 2
        elif self.annot1[tag_type][sent] == self.annot2[tag_type][sent]:
            ws.cell(row=row, column=col).style = '40 % - Accent6'
            ws.cell(row=row, column=col+3).style = '40 % - Accent6'
            ws.cell(row=row, column=col+6).style = 'Bad'
            ws.cell(row=row, column=col+9).style = 'Bad'
            self.agree[tag_type] += 1
        elif self.annot1[tag_type][sent] == self.annot3[tag_type][sent]:
            ws.cell(row=row, column=col).style = '40 % - Accent6'
            ws.cell(row=row, column=col+3).style = 'Bad'
            ws.cell(row=row, column=col+6).style = '40 % - Accent6'
            ws.cell(row=row, column=col+9).style = 'Bad'
            self.agree[tag_type] += 1
        elif self.annot1[tag_type][sent] == self.annot4[tag_type][sent]:
            ws.cell(row=row, column=col).style = '40 % - Accent6'
            ws.cell(row=row, column=col+3).style = 'Bad'
            ws.cell(row=row, column=col+6).style = 'Bad'
            ws.cell(row=row, column=col+9).style = '40 % - Accent6'
            self.agree[tag_type] += 1
        elif self.annot2[tag_type][sent] == self.annot3[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Bad'
            ws.cell(row=row, column=col+3).style = '40 % - Accent6'
            ws.cell(row=row, column=col+6).style = '40 % - Accent6'
            ws.cell(row=row, column=col+9).style = 'Bad'
            self.agree[tag_type] += 1
        elif self.annot2[tag_type][sent] == self.annot4[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Bad'
            ws.cell(row=row, column=col+3).style = '40 % - Accent6'
            ws.cell(row=row, column=col+6).style = 'Bad'
            ws.cell(row=row, column=col+9).style = '40 % - Accent6'
            self.agree[tag_type] += 1
        elif self.annot3[tag_type][sent] == self.annot4[tag_type][sent]:
            ws.cell(row=row, column=col).style = 'Bad'
            ws.cell(row=row, column=col+3).style = 'Bad'
            ws.cell(row=row, column=col+6).style = '40 % - Accent6'
            ws.cell(row=row, column=col+9).style = '40 % - Accent6'
            self.agree[tag_type] += 1
        else:
            ws.cell(row=row, column=col).style = 'Bad'
            ws.cell(row=row, column=col+3).style = 'Bad'
            ws.cell(row=row, column=col+6).style = 'Bad'
            ws.cell(row=row, column=col+9).style = 'Bad'
    def pair_helper(self, tag_type, sent, ws, factor):
        """
        determines how much agreement there is and colors the attribute accordingly
        """
        if self.annot1[tag_type][sent] == self.annot2[tag_type][sent] == self.annot3[tag_type][sent] == self.annot4[tag_type][sent]:
            self.agree['main'] += 4 * factor
        elif self.annot1[tag_type][sent] == self.annot2[tag_type][sent] == self.annot3[tag_type][sent]:
            self.agree['main'] += 3 * factor
        elif self.annot1[tag_type][sent] == self.annot2[tag_type][sent] == self.annot4[tag_type][sent]:
            self.agree['main'] += 3 * factor
        elif self.annot1[tag_type][sent] == self.annot3[tag_type][sent] == self.annot4[tag_type][sent]:
            self.agree['main'] += 3 * factor
        elif self.annot2[tag_type][sent] == self.annot3[tag_type][sent] == self.annot4[tag_type][sent]:
            self.agree['main'] += 3 * factor
        elif self.annot1[tag_type][sent] == self.annot2[tag_type][sent] and self.annot3[tag_type][sent] == self.annot4[tag_type][sent]:
            self.agree['main'] += 2 * factor
        elif self.annot1[tag_type][sent] == self.annot3[tag_type][sent] and self.annot2[tag_type][sent] == self.annot4[tag_type][sent]:
            self.agree['main'] += 2 * factor
        elif self.annot1[tag_type][sent] == self.annot4[tag_type][sent] and self.annot2[tag_type][sent] == self.annot3[tag_type][sent]:
            self.agree['main'] += 2 * factor
        elif self.annot1[tag_type][sent] == self.annot2[tag_type][sent]:
            self.agree['main'] += 1 * factor
        elif self.annot1[tag_type][sent] == self.annot3[tag_type][sent]:
            self.agree['main'] += 1 * factor
        elif self.annot1[tag_type][sent] == self.annot4[tag_type][sent]:
            self.agree['main'] += 1 * factor
        elif self.annot2[tag_type][sent] == self.annot3[tag_type][sent]:
            self.agree['main'] += 1 * factor
        elif self.annot2[tag_type][sent] == self.annot4[tag_type][sent]:
            self.agree['main'] += 1 * factor
        elif self.annot3[tag_type][sent] == self.annot4[tag_type][sent]:
            self.agree['main'] += 1 * factor
    def stats(self, file_name):
        ws = self.wb.create_sheet("tables")
        self.table_1_2(ws)
        self.table_1_2_all(ws)
        ws.cell(row=12, column=1, value='NARRATION_AND_DESCRIPTION')
        ws.cell(row=13, column=1, value='ABOUT_THE_GAME')
        ws.cell(row=14, column=1, value='MECHANICS')
        ws.cell(row=15, column=1, value='NON-GAME_RELATED')
        ws.cell(row=16, column=1, value='NON-CONTENT')
        ws.cell(row=11, column=2, value='annot 1')
        self.tag_counter(ws, 12, 2, 'B')
        ws.cell(row=11, column=3, value='annot 2')
        self.tag_counter(ws, 12, 3, 'E')
        ws.cell(row=11, column=4, value='annot 3')
        self.tag_counter(ws, 12, 4, 'H')
        ws.cell(row=11, column=5, value='annot 4')
        self.tag_counter(ws, 12, 5, 'K')
        ws.cell(row=11, column=6, value='sum')
        rows = 12
        for tag in self.main_tags:
            ws.cell(row=rows, column=6, value='=sum(B'+str(rows)+':E'+str(rows)+')')
            rows += 1
        ws.cell(row=11, column=7, value='P(tag)')
        rows = 12
        for tag in self.main_tags:
            ws.cell(row=rows, column=7, value='=F'+str(rows)+'/'+str(4*len(self.hum_sent_nums)))
            rows += 1
        ws.cell(row=17, column=5, value='chance:')
        ws.cell(row=17, column=6, value='=G12^2+G13^2+G14^2+G15^2+G16^2')
        ws.cell(row=18, column=5, value='observed')
        ws.cell(row=18, column=6, value=self.agree['main tag']/(4*len(self.hum_sent_nums)))
        ws.cell(row=19, column=5, value='κ')
        ws.cell(row=19, column=6, value='=(F18-F17)/(1-F17)')
        
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=10, value=tag)
            rows += 1
        ws.cell(row=20, column=11, value='annot 1')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=11, value=self.annot1_freq[tag])
            rows += 1
        ws.cell(row=20, column=12, value='annot 2')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=12, value=self.annot2_freq[tag])
            rows += 1
        ws.cell(row=20, column=13, value='annot 3')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=13, value=self.annot3_freq[tag])
            rows += 1
        ws.cell(row=20, column=14, value='annot 4')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=14, value=self.annot4_freq[tag])
            rows += 1
        ws.cell(row=20, column=15, value='sum')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=15, value='=sum(K'+str(rows)+':N'+str(rows)+')')
            rows += 1
        ws.cell(row=20, column=16, value='P(tag)')
        rows = 21
        for tag in self.all_tags:
            ws.cell(row=rows, column=16, value='=O'+str(rows)+'/'+str(4*len(self.hum_sent_nums)))
            rows += 1
        ws.cell(row=32, column=15, value='chance:')
        ws.cell(row=32, column=16, value='=P21^2+P22^2+P23^2+P24^2+P25^2+P26^2+P27^2+P28^2+P29^2+P30^2+P31^2')
        ws.cell(row=33, column=15, value='observed')
        ws.cell(row=33, column=16, value=self.agree['main']/(4*len(self.hum_sent_nums)))
        ws.cell(row=34, column=15, value='κ')
        ws.cell(row=34, column=16, value='=(P33-P32)/(1-P32)')
        self.wb.save(file_name)
        
if __name__ == '__main__':
#    thingy30 = gold()
#    thingy30.find_data1('30_goldstandard.xml')
#    thingy30.find_data2('30_alex1.xml')
#    thingy30.find_data3('30_Kirsten.xml')
#    thingy30.print_to_excel('30_3.xlsx')
#    thingy30.stats('30_3.xlsx')
#    
#    thingy31 = gold()
#    thingy31.find_data1('31_goldstandard.xml')
#    thingy31.find_data2('31_Bingyang.xml')
#    thingy31.find_data3('31_danielle.xml')
#    thingy31.print_to_excel('31_3.xlsx')
#    thingy31.stats('31_3.xlsx')
#    
#    thingy32 = gold()
#    thingy32.find_data1('32_goldstandard.xml')
#    thingy32.find_data2('32_alex.xml')
#    thingy32.find_data3('32_Bingyang.xml')
#    thingy32.print_to_excel('32_3.xlsx')
#    thingy32.stats('32_3.xlsx')
#    
#    thingy33 = gold()
#    thingy33.find_data1('33_goldstandard.xml')
#    thingy33.find_data2('33_danielle.xml')
#    thingy33.find_data3('33_Kirsten.xml')
#    thingy33.print_to_excel('33_3.xlsx')
#    thingy33.stats('33_3.xlsx')
    
    thingy34 = gold()
    thingy34.find_data1('34_goldstandard.xml')
    thingy34.find_data2('34_alex.xml')
    thingy34.find_data3('34_danielle.xml')
    thingy34.print_to_excel('test.xlsx')
    thingy34.stats('test.xlsx')
    
#    thingy35 = gold()
#    thingy35.find_data1('35_goldstandard.xml')
#    thingy35.find_data2('35_Bingyang.xml')
#    thingy35.find_data3('35_Kirsten.xml')
#    thingy35.print_to_excel('35_3.xlsx')
#    thingy35.stats('35_3.xlsx')
#    
#    thingy36 = gold()
#    thingy36.find_data1('36_goldstandard.xml')
#    thingy36.find_data2('36_alex.xml')
#    thingy36.find_data3('36_Kirsten.xml')
#    thingy36.print_to_excel('36_3.xlsx')
#    thingy36.stats('36_3.xlsx')
#    
#    thingy37 = gold()
#    thingy37.find_data1('37_goldstandard.xml')
#    thingy37.find_data2('37_Bingyang.xml')
#    thingy37.find_data3('37_danielle.xml')
#    thingy37.print_to_excel('37_3.xlsx')
#    thingy37.stats('37_3.xlsx')
#    